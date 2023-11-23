import re
import pandas as pd
from collections import namedtuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

"""

主要用于将  mvn dependence:tree 输出的依赖项进行梳理，写入excel
"""

pattern1 = r'\[[^\]]+\]'
pattern2 = r"\+\-\s(.+)"
# 添加过滤规则
filter_rule = ['cn.sunline', 'SNAPSHOT']


def comb_file(file):
    content = []
    with open(file, "r") as f:
        lines = f.readlines()
    for line in lines:
        if ':jar:' in line and all(rule not in line for rule in filter_rule):
            # 获取到携带依赖的行内容
            new_line = re.sub(pattern1, '', line)
            match = re.search(pattern2, new_line)
            if match:
                content.append(match.group(1))
    to_excel_auto_title_len(list(set(content)))


Dependency = namedtuple('Dependency', ['group_id', 'artifact_id', 'package', 'version'])


def parse_dependency(line):
    # line format is "groupId:artifactId:packaging:version:scope"
    parts = line.split(':')
    return Dependency(*parts[0:4])


def to_excel_auto_title_len(content):
    dependencies = [parse_dependency(line.strip()) for line in content]
    df = pd.DataFrame(dependencies, columns=['Group ID', 'Artifact ID', 'Package', 'Version'])
    # 创建一个新的 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    # 写入表头
    header = df.columns.tolist()
    ws.append(header)
    # 将 DataFrame 数据转换为行列表
    data_rows = df.values.tolist()

    # 写入数据行到工作表
    for row in data_rows:
        ws.append(row)
    # 自动调整每列的宽度
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # 可根据需要进行微调
        ws.column_dimensions[column_letter].width = adjusted_width
    # 保存 Excel 文件
    wb.save('dependencies.xlsx')


def to_excel_no_auto_title_len(content):
    dependencies = [parse_dependency(line.strip()) for line in content]
    df = pd.DataFrame(dependencies, columns=['Group ID', 'Artifact ID', 'Package', 'Version'])
    df.to_excel('dependencies.xlsx')


def run(file):
    comb_file(file)


if __name__ == '__main__':
    """
        待处理文件
    """
    run("/Users/lsr/Desktop/maventree.txt")
